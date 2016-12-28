#Code used to test for statistically significant differences in symptoms between different diseases.

from openpyxl import load_workbook
import numpy as np
from scipy import stats



def Significance(set1,set2):
    SE = ( (np.std(set1)**2/len(set1))+ (np.std(set2)**2/len(set2)) )
    T = (  (np.mean(set1)-np.mean(set2))-0)/SE
    if len(set1)> len(set2): F=len(set2)-1
    else: F= len(set1)-1

    return stats.t.sf(T,df=F)


feverxl = load_workbook(filename='fever.xlsx', read_only=True)
feverws = feverxl['fever'] # ws is now an IterableWorksheet

WBCxl = load_workbook(filename='WBC.xlsx', read_only=True)
WBCws = WBCxl['WBC'] # ws is now an IterableWorksheet


#Create lists for data

NoFever= []
NoWBC = []
Py2Fever = []
Py2WBC =[]
Py3Fever = []
Py3WBC = []

#P of someone having fever above 102
p102=0.0
pSickAnd102 = 0.0
pSick = 0.0

for row in feverws.rows:
    if(row[0].value)== "No": NoFever.append(float(row[1].value))

    if(row[0].value)== "Py2.7":
        Py2Fever.append(float(row[1].value))

    if(row[0].value)== "Py3":
        Py3Fever.append(float(row[1].value))


    if(row[1].value>=102): p102+=1
    if(row[0].value)== "Py2.7" or (row[0].value)== "Py3" :
        pSick+=1
        if(row[1].value>=102): pSickAnd102+=1

pSick=pSick/89.0
p102=p102/89.0
pSickAnd102=pSickAnd102/89.0

for row in WBCws.rows:
    if(row[0].value)== "No": NoWBC.append(float(row[1].value))
    if(row[0].value)== "Py2.7": Py2WBC.append(float(row[1].value))
    if(row[0].value)== "Py3": Py3WBC.append(float(row[1].value))



#Use this for P(102 given D)
p102GivenSick=0.0
for x in Py2Fever:
    if x>=102: p102GivenSick+=1
for x in Py3Fever:
    if x>=102: p102GivenSick+=1
    #Divide by number of sick people
p102GivenSick=p102GivenSick/39.0





#for table
print Significance(Py2Fever,Py3Fever)
print Significance(Py2WBC,Py3WBC)
print Significance(Py2Fever,NoFever)
print Significance(Py2WBC,NoWBC)
print Significance(Py3Fever,NoFever)
print Significance(Py3WBC,NoWBC)

#temperatures
print np.mean(NoFever)
print np.mean(Py2Fever)
print np.mean(Py3Fever)

#Question 2
print pSickAnd102 / p102

print p102GivenSick*pSick/p102
print pSick

